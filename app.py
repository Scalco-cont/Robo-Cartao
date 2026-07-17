import os
import uuid
import sqlite3
import threading
from datetime import datetime
from flask import Flask, request, jsonify, send_file, session, render_template
from flask_cors import CORS
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import io

from config import Config, PARAMETROS, MESSAGES

app = Flask(__name__)
app.secret_key = Config.SECRET_KEY
CORS(app)

# Configurações
os.makedirs(Config.UPLOAD_FOLDER, exist_ok=True)
os.makedirs(Config.OUTPUT_FOLDER, exist_ok=True)

# Armazenamento de sessões: SQLite compartilhado entre os workers gunicorn.
# Sem volume — o arquivo vive no filesystem do container e some a cada redeploy,
# igual ao comportamento anterior em memória (decisão registrada em DESIGN_1.1_1.3.md).
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'sessions.db')


def get_conn():
    conn = sqlite3.connect(DB_PATH, timeout=10, isolation_level=None)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=10000")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db():
    conn = get_conn()
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sessions (
                session_id TEXT PRIMARY KEY,
                processing INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS files (
                file_id TEXT PRIMARY KEY,
                session_id TEXT NOT NULL REFERENCES sessions(session_id),
                filename TEXT NOT NULL,
                original_filename TEXT NOT NULL,
                filepath TEXT NOT NULL,
                status TEXT NOT NULL,
                error TEXT,
                txt_content TEXT,
                txt_filename TEXT,
                txt_path TEXT,
                upload_time TEXT NOT NULL
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_files_session ON files(session_id)")
    finally:
        conn.close()


init_db()


def allowed_file(filename):
    """Verifica se o arquivo tem extensão permitida"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in Config.ALLOWED_EXTENSIONS


def normalizar(texto):
    """Normaliza texto para comparação"""
    if texto is None:
        return ""
    return (
        str(texto)
        .replace("\n", " ")
        .replace("\r", " ")
        .strip()
        .lower()
    )


def formatar_data(data_value):
    """
    Formata a data para o padrão dd/mm/yyyy
    Aceita diversos formatos de entrada e remove hora se existir
    """
    from datetime import datetime
    import re
    
    if data_value is None:
        return ""
    
    # Se for um objeto datetime do Python/Excel
    if hasattr(data_value, "strftime"):
        return data_value.strftime("%d/%m/%Y")
    
    # Se for string
    data_str = str(data_value).strip()
    
    # Dicionário para converter meses em português (abreviados e completos)
    meses_pt = {
        'jan': '01', 'janeiro': '01',
        'fev': '02', 'fevereiro': '02',
        'mar': '03', 'março': '03', 'marco': '03',
        'abr': '04', 'abril': '04',
        'mai': '05', 'maio': '05',
        'jun': '06', 'junho': '06',
        'jul': '07', 'julho': '07',
        'ago': '08', 'agosto': '08',
        'set': '09', 'setembro': '09',
        'out': '10', 'outubro': '10',
        'nov': '11', 'novembro': '11',
        'dez': '12', 'dezembro': '12'
    }
    
    # Verifica se é um formato com nome de mês em português (ex: "1 de ago. de 2025 07:45")
    # Padrão: número + "de" + mês + "de" + ano
    padrao_mes_pt = r'(\d{1,2})\s+de\s+([a-zç]+)\.?\s+de\s+(\d{4})'
    match = re.search(padrao_mes_pt, data_str.lower())
    
    if match:
        dia = match.group(1).zfill(2)  # Adiciona zero à esquerda se necessário
        mes_texto = match.group(2).replace('.', '').strip()
        ano = match.group(3)
        
        # Busca o mês no dicionário
        mes_numero = meses_pt.get(mes_texto)
        
        if mes_numero:
            return f"{dia}/{mes_numero}/{ano}"
    
    # Se contém espaço (provavelmente tem hora), pega só a parte da data
    if ' ' in data_str:
        data_str = data_str.split(' ')[0]
    
    # Tenta converter diferentes formatos de string para datetime
    formatos_possiveis = [
        "%d/%m/%Y",      # 19/12/2025
        "%Y-%m-%d",      # 2025-12-19
        "%d-%m-%Y",      # 19-12-2025
        "%m/%d/%Y",      # 12/19/2025
        "%d/%m/%y",      # 19/12/25
        "%Y/%m/%d",      # 2025/12/19
    ]
    
    for formato in formatos_possiveis:
        try:
            data_obj = datetime.strptime(data_str, formato)
            return data_obj.strftime("%d/%m/%Y")
        except ValueError:
            continue
    
    # Se nenhum formato funcionou, retorna a string original sem a hora
    return data_str


def formatar_numero(valor, remover_negativo=False):
    """
    Formata números mantendo vírgula como separador decimal
    e garantindo sempre 2 casas decimais.
    Se remover_negativo for True, retorna o valor absoluto.
    """
    if valor is None or valor == "":
        return ""
    
    # Se for número (int ou float)
    if isinstance(valor, (int, float)):
        num = float(valor)
        if remover_negativo:
            num = abs(num)
        # Formata com 2 casas decimais e substitui ponto por vírgula
        valor_formatado = "{:.2f}".format(num)
        return valor_formatado.replace('.', ',')
    
    # Se for string, tenta converter para float primeiro
    valor_str = str(valor).strip()
    
    # Remove símbolos de moeda (R$, $) e espaços
    valor_str = valor_str.replace('R$', '').replace('$', '').replace(' ', '')
    
    # Se está vazio após limpeza
    if not valor_str:
        return ""
    
    try:
        # ⭐ CORREÇÃO: Identifica o separador decimal e remove os separadores de milhar
        # Exemplos problemáticos:
        # "26,900,000000" -> múltiplas vírgulas (Stone)
        # "26.900,00" -> ponto de milhar, vírgula decimal (Brasil)
        # "1,234.56" -> vírgula de milhar, ponto decimal (EUA)
        
        # Conta quantos pontos e vírgulas tem
        num_virgulas = valor_str.count(',')
        num_pontos = valor_str.count('.')
        
        # Se tem múltiplas vírgulas OU múltiplos pontos, é separador de milhar
        if num_virgulas > 1 and num_pontos == 0:
            # Múltiplas vírgulas SEM pontos: "26,900,000" -> a ÚLTIMA vírgula é o decimal
            ultima_virgula = valor_str.rfind(',')
            antes = valor_str[:ultima_virgula].replace(',', '')
            depois = valor_str[ultima_virgula+1:]
            valor_str = antes + '.' + depois
        elif num_pontos > 1 and num_virgulas == 0:
            # Múltiplos pontos SEM vírgulas: "1.234.567" -> o ÚLTIMO ponto é o decimal
            ultimo_ponto = valor_str.rfind('.')
            antes = valor_str[:ultimo_ponto].replace('.', '')
            depois = valor_str[ultimo_ponto+1:]
            valor_str = antes + '.' + depois
        elif num_pontos > 1 and num_virgulas == 1:
            # Múltiplos pontos COM uma vírgula: "1.234.567,89" -> vírgula é decimal, pontos são milhares
            valor_str = valor_str.replace('.', '').replace(',', '.')
        elif num_virgulas > 1 and num_pontos == 1:
            # Múltiplas vírgulas COM um ponto: "1,234,567.89" -> ponto é decimal, vírgulas são milhares
            valor_str = valor_str.replace(',', '')
        elif num_virgulas == 1 and num_pontos == 1:
            # Tem UMA vírgula E UM ponto: precisa identificar qual é o decimal
            # O decimal vem por último
            if valor_str.rindex(',') > valor_str.rindex('.'):
                # Vírgula vem depois: "1.234,56" (formato BR)
                valor_str = valor_str.replace('.', '').replace(',', '.')
            else:
                # Ponto vem depois: "1,234.56" (formato US)
                valor_str = valor_str.replace(',', '')
        elif num_virgulas == 1:
            # Apenas UMA vírgula: é o separador decimal
            valor_str = valor_str.replace(',', '.')
        # Se tem apenas ponto(s), já está no formato correto
        
        valor_num = float(valor_str)
        if remover_negativo:
            valor_num = abs(valor_num)
        # Formata com 2 casas decimais e substitui ponto por vírgula
        valor_formatado = "{:.2f}".format(valor_num)
        return valor_formatado.replace('.', ',')
    except ValueError:
        # Se não conseguir converter, retorna como string
        # mas tenta garantir vírgula como separador
        if '.' in valor_str:
            return valor_str.replace('.', ',')
        return valor_str


def encontrar_coluna_tipo(ws, nome_coluna):
    """
    Encontra a coluna que identifica o tipo (débito/crédito).
    Retorna o número da coluna ou None se não encontrar.
    """
    nome_normalizado = normalizar(nome_coluna)
    
    # Procura nas primeiras 20 linhas
    for row in ws.iter_rows(max_row=20):
        for cell in row:
            if not cell.value:
                continue
            
            if normalizar(cell.value) == nome_normalizado:
                return cell.column
    
    return None


def identificar_tipo_da_linha(ws, linha, col_tipo, config_identificacao):
    """
    Identifica se uma linha específica é débito ou crédito.
    Retorna 'debito', 'credito' ou None.
    """
    valor_celula = ws.cell(row=linha, column=col_tipo).value
    
    if not valor_celula:
        return None
    
    valor_normalizado = normalizar(valor_celula)
    
    # Verifica se é débito
    for valor_debito in config_identificacao['debito']['valores']:
        if normalizar(valor_debito) in valor_normalizado:
            return 'debito'
    
    # Verifica se é crédito
    for valor_credito in config_identificacao['credito']['valores']:
        if normalizar(valor_credito) in valor_normalizado:
            return 'credito'
    
    return None


def carregar_excel_robusto(caminho_xlsx):
    """
    Tenta carregar o arquivo Excel usando diferentes métodos.
    Retorna (workbook, mensagem_erro) onde mensagem_erro é None em caso de sucesso.
    """
    try:
        # Método 1: Tenta carregar normalmente
        wb = load_workbook(caminho_xlsx)
        return wb, None
    except Exception as e1:
        error_msg = str(e1).lower()
        
        # Se o erro for relacionado a XML/stylesheet inválido
        if 'stylesheet' in error_msg or 'xml' in error_msg or 'invalid' in error_msg:
            try:
                # Método 2: Tenta carregar sem estilos (data_only=True)
                wb = load_workbook(caminho_xlsx, data_only=True)
                return wb, None
            except Exception as e2:
                try:
                    # Método 3: Tenta usar pandas para ler e depois salvar novamente
                    import pandas as pd
                    import tempfile
                    
                    # Lê o arquivo com pandas (mais robusto para arquivos corrompidos)
                    df = pd.read_excel(caminho_xlsx, engine='openpyxl')
                    
                    # Salva em um arquivo temporário limpo
                    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
                        temp_path = tmp.name
                        df.to_excel(temp_path, index=False, engine='openpyxl')
                    
                    # Tenta carregar o arquivo limpo
                    wb = load_workbook(temp_path)
                    
                    # Remove arquivo temporário
                    try:
                        os.remove(temp_path)
                    except:
                        pass
                    
                    return wb, None
                    
                except Exception as e3:
                    try:
                        # Método 4: Tenta com xlrd (para arquivos .xls antigos salvos como .xlsx)
                        import pandas as pd
                        import tempfile
                        
                        # Tenta ler como XLS
                        df = pd.read_excel(caminho_xlsx, engine='xlrd')
                        
                        # Salva como XLSX limpo
                        with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
                            temp_path = tmp.name
                            df.to_excel(temp_path, index=False, engine='openpyxl')
                        
                        wb = load_workbook(temp_path)
                        
                        try:
                            os.remove(temp_path)
                        except:
                            pass
                        
                        return wb, None
                    except Exception as e4:
                        try:
                            # Método 5: EXTRAÇÃO MANUAL DO XML (último recurso)
                            # Este método funciona mesmo quando o stylesheet está totalmente corrompido
                            import zipfile
                            import xml.etree.ElementTree as ET
                            import pandas as pd
                            import tempfile
                            
                            # Namespace do Excel
                            ns = {'': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                            
                            # Abre o XLSX como ZIP e extrai os dados
                            with zipfile.ZipFile(caminho_xlsx, 'r') as zip_ref:
                                # Lê as strings compartilhadas
                                try:
                                    with zip_ref.open('xl/sharedStrings.xml') as f:
                                        shared_strings_root = ET.fromstring(f.read())
                                        shared_strings = []
                                        for si in shared_strings_root.findall('.//t', namespaces=ns):
                                            shared_strings.append(si.text if si.text else '')
                                except:
                                    shared_strings = []
                                
                                # Lê a planilha principal
                                with zip_ref.open('xl/worksheets/sheet1.xml') as f:
                                    sheet_root = ET.fromstring(f.read())
                                
                                # Extrai as linhas
                                rows_data = []
                                for row in sheet_root.findall('.//row', namespaces=ns):
                                    row_data = {}
                                    for cell in row.findall('.//c', namespaces=ns):
                                        ref = cell.get('r')
                                        col_letter = ''.join([c for c in ref if c.isalpha()])
                                        cell_type = cell.get('t')
                                        v_elem = cell.find('.//v', namespaces=ns)
                                        
                                        if v_elem is not None:
                                            value = v_elem.text
                                            # Se for string compartilhada, busca no array
                                            if cell_type == 's' and shared_strings:
                                                try:
                                                    value = shared_strings[int(value)]
                                                except:
                                                    pass
                                            row_data[col_letter] = value
                                    
                                    if row_data:
                                        rows_data.append(row_data)
                                
                                # Encontra o cabeçalho (procura por colunas típicas)
                                # IMPORTANTE: procura por uma linha que tenha MÚLTIPLAS palavras-chave
                                # para evitar pegar linhas de resumo (como "Valor bruto: R$ 123,45")
                                header_keywords = ['data', 'valor', 'taxa', 'produto', 'cartão', 'cartões']
                                header_row_idx = None
                                
                                for i, row in enumerate(rows_data):
                                    # Conta quantas keywords aparecem nesta linha
                                    row_values_lower = [str(v).lower() for v in row.values()]
                                    row_text = ' '.join(row_values_lower)
                                    keyword_count = sum(1 for keyword in header_keywords if keyword in row_text)
                                    
                                    # Precisa ter pelo menos 3 keywords E mais de 5 colunas
                                    # Isso garante que não pegamos linhas de resumo com uma única célula
                                    if keyword_count >= 3 and len(row) >= 5:
                                        header_row_idx = i
                                        break
                                
                                if header_row_idx is None:
                                    raise Exception("Cabeçalho não encontrado na planilha")
                                
                                # Cria DataFrame
                                headers = list(rows_data[header_row_idx].values())
                                data_rows = rows_data[header_row_idx + 1:]
                                
                                df_data = []
                                for row in data_rows:
                                    row_list = [row.get(col, '') for col in rows_data[header_row_idx].keys()]
                                    df_data.append(row_list)
                                
                                df = pd.DataFrame(df_data, columns=headers)
                                
                                # Salva em arquivo temporário
                                with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
                                    temp_path = tmp.name
                                    df.to_excel(temp_path, index=False, engine='openpyxl')
                                
                                # Carrega o arquivo reconstruído
                                wb = load_workbook(temp_path)
                                
                                try:
                                    os.remove(temp_path)
                                except:
                                    pass
                                
                                return wb, None
                                
                        except Exception as e5:
                            # Se TODOS os métodos falharem, retorna erro detalhado
                            erro_detalhado = (
                                f"❌ Não foi possível abrir o arquivo Excel após tentar 5 métodos diferentes.\n\n"
                                f"Métodos tentados:\n"
                                f"1. OpenPyXL padrão\n"
                                f"2. OpenPyXL data_only\n"
                                f"3. Pandas com openpyxl\n"
                                f"4. Pandas com xlrd\n"
                                f"5. Extração manual do XML\n\n"
                                f"Possíveis causas:\n"
                                f"• Arquivo severamente corrompido\n"
                                f"• Arquivo protegido por senha\n"
                                f"• Formato não suportado\n\n"
                                f"Soluções:\n"
                                f"• Abra o arquivo no Excel e salve novamente\n"
                                f"• Exporte para CSV e reimporte para Excel\n\n"
                                f"Erro original: {str(e1)}\n"
                                f"Último erro: {str(e5)}"
                            )
                            return None, erro_detalhado
        else:
            # Erro não relacionado a XML/stylesheet
            return None, f"Erro ao abrir arquivo Excel: {str(e1)}"


def get_session_id():
    """Obtém ou cria um ID de sessão único"""
    if 'session_id' not in session:
        session['session_id'] = str(uuid.uuid4())
    return session['session_id']


def init_session_data(session_id):
    """Garante que a sessão existe no banco compartilhado entre workers"""
    conn = get_conn()
    try:
        conn.execute(
            "INSERT OR IGNORE INTO sessions (session_id, processing, created_at) VALUES (?, 0, ?)",
            (session_id, datetime.now().isoformat())
        )
    finally:
        conn.close()


def try_start_processing(session_id):
    """(True, None) se marcou processing=1; (False, motivo) se não pôde.
    Propaga sqlite3.OperationalError se não conseguir nem abrir a transação (lock ocupado)."""
    conn = get_conn()

    try:
        conn.execute("BEGIN IMMEDIATE")
    except sqlite3.OperationalError:
        conn.close()
        raise  # sem transação aberta — não tentar ROLLBACK aqui

    try:
        row = conn.execute("SELECT processing FROM sessions WHERE session_id=?", (session_id,)).fetchone()
        if row and row[0]:
            conn.execute("ROLLBACK")
            return False, "Já existe um processamento em andamento"

        n = conn.execute(
            "SELECT COUNT(*) FROM files WHERE session_id=? AND status='uploaded'", (session_id,)
        ).fetchone()[0]
        if n == 0:
            conn.execute("ROLLBACK")
            return False, "Nenhum arquivo para processar"

        conn.execute("UPDATE sessions SET processing=1 WHERE session_id=?", (session_id,))
        conn.execute("COMMIT")
        return True, None
    except Exception:
        conn.execute("ROLLBACK")
        raise
    finally:
        conn.close()


def set_processing(session_id, value):
    conn = get_conn()
    try:
        conn.execute("UPDATE sessions SET processing=? WHERE session_id=?", (value, session_id))
    finally:
        conn.close()


def cleanup_completed_files(session_id):
    """Remove do disco (xlsx + txt) e do banco os arquivos já concluídos/com erro."""
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT filepath, txt_path FROM files WHERE session_id=? AND status IN ('completed','error')",
            (session_id,)
        ).fetchall()
        for filepath, txt_path in rows:
            for p in (filepath, txt_path):
                if p and os.path.exists(p):
                    try:
                        os.remove(p)
                    except Exception:
                        pass
        conn.execute(
            "DELETE FROM files WHERE session_id=? AND status IN ('completed','error')", (session_id,)
        )
        return len(rows)
    finally:
        conn.close()


def process_xlsx_to_txt(caminho_xlsx, nome_arquivo):
    """
    Processa um arquivo XLSX e retorna o conteúdo TXT como string
    CORRIGIDO: Procura as 3 colunas (data, valor, taxa) na MESMA LINHA
    """
    nome_base = os.path.splitext(nome_arquivo)[0]
    nome_lower = nome_base.lower()

    # Identifica o tipo pelo nome
    parametro = None
    tipo_identificado = None
    for chave in PARAMETROS:
        if chave in nome_lower:
            parametro = PARAMETROS[chave]
            tipo_identificado = chave
            break

    if not parametro:
        return None, MESSAGES['type_not_identified']

    try:
        # Usa o carregador robusto
        wb, erro = carregar_excel_robusto(caminho_xlsx)
        
        if wb is None:
            return None, erro
        
        ws = wb.active

        # Verifica se precisa identificar débito/crédito por coluna
        usar_identificacao_por_coluna = "identificar_por_coluna" in parametro
        col_tipo = None
        
        # Validação: se não usar identificação por coluna, deve ter linha1 e linha2
        if not usar_identificacao_por_coluna:
            if 'linha1' not in parametro or 'linha2' not in parametro:
                return None, "❌ Configuração incorreta: faltam 'linha1' e 'linha2' no config.py"
        else:
            # Se usar identificação por coluna, valida a configuração
            config_id = parametro['identificar_por_coluna']
            if 'nome_coluna' not in config_id:
                return None, "❌ Configuração incorreta: falta 'nome_coluna' em 'identificar_por_coluna'"
            
            # Procura a coluna de tipo
            col_tipo = encontrar_coluna_tipo(ws, config_id['nome_coluna'])
            if col_tipo is None:
                return None, MESSAGES['type_column_not_found'].format(column=config_id['nome_coluna'])

        # NOVA LÓGICA: Procura a linha que contém AS 3 COLUNAS juntas
        palavra_data = parametro.get("palavra_data")
        palavra_valor = parametro.get("palavra_valor")
        palavra_taxa = parametro.get("palavra_taxa")

        col_data = None
        col_valor = None
        col_taxa = None
        linha_inicio = None

        # Busca a linha que tem as 3 colunas
        for row_idx, row in enumerate(ws.iter_rows(max_row=50), start=1):
            # Armazena as colunas encontradas nesta linha
            encontrou_data_nesta_linha = None
            encontrou_valor_nesta_linha = None
            encontrou_taxa_nesta_linha = None
            
            for cell in row:
                if not cell.value:
                    continue

                valor_normalizado = normalizar(cell.value)

                # Data
                if normalizar(palavra_data) == valor_normalizado:
                    encontrou_data_nesta_linha = cell.column

                # Valor
                if normalizar(palavra_valor) == valor_normalizado:
                    encontrou_valor_nesta_linha = cell.column

                # Taxa
                if normalizar(palavra_taxa) == valor_normalizado:
                    encontrou_taxa_nesta_linha = cell.column
            
            # Se encontrou as 3 colunas na mesma linha, usa essa linha!
            if encontrou_data_nesta_linha and encontrou_valor_nesta_linha and encontrou_taxa_nesta_linha:
                col_data = encontrou_data_nesta_linha
                col_valor = encontrou_valor_nesta_linha
                col_taxa = encontrou_taxa_nesta_linha
                linha_inicio = row_idx
                break

        # Validação das colunas encontradas
        erros = []
        if not col_data:
            erros.append(MESSAGES['column_not_found'].format(column="Data", word=palavra_data))
        if not col_valor:
            erros.append(MESSAGES['column_not_found'].format(column="Valor", word=palavra_valor))
        if not col_taxa:
            erros.append(MESSAGES['column_not_found'].format(column="Taxa", word=palavra_taxa))

        if erros:
            return None, "\n".join(erros)

        # Processa os dados linha por linha
        linhas_saida = []
        linha_atual = linha_inicio + 1

        while True:
            data_cell = ws.cell(row=linha_atual, column=col_data)
            valor_cell = ws.cell(row=linha_atual, column=col_valor)
            taxa_cell = ws.cell(row=linha_atual, column=col_taxa)

            # Se todas as células estão vazias, termina
            if not data_cell.value and not valor_cell.value and not taxa_cell.value:
                break

            # Se pelo menos uma das células principais tem valor, processa
            if data_cell.value or valor_cell.value:
                # ⭐ NOVO: Verifica se deve ignorar valores negativos
                if parametro.get("ignorar_valores_negativos", False):
                    # Se o valor bruto for negativo, pula esta linha
                    if valor_cell.value is not None:
                        try:
                            valor_num = float(str(valor_cell.value).replace(',', '.').replace(' ', ''))
                            if valor_num < 0:
                                linha_atual += 1
                                continue  # Pula para a próxima linha
                        except (ValueError, AttributeError):
                            pass  # Se não conseguir converter, continua normalmente
                
                data_formatada = formatar_data(data_cell.value)
                valor_formatado = formatar_numero(valor_cell.value)
                # Remove o sinal negativo da taxa
                taxa_formatada = formatar_numero(taxa_cell.value, remover_negativo=True)

                # Determina qual linha1 e linha2 usar
                if usar_identificacao_por_coluna:
                    # Identifica o tipo desta linha específica
                    tipo_linha = identificar_tipo_da_linha(
                        ws, linha_atual, col_tipo, 
                        parametro['identificar_por_coluna']
                    )
                    
                    if tipo_linha:
                        linha1 = parametro['identificar_por_coluna'][tipo_linha]['linha1']
                        linha2 = parametro['identificar_por_coluna'][tipo_linha]['linha2']
                    else:
                        # Se não identificar o tipo, pula a linha
                        linha_atual += 1
                        continue
                else:
                    # Usa linha1 e linha2 fixos do parametro
                    linha1 = parametro['linha1']
                    linha2 = parametro['linha2']

                # Formata as linhas de saída na ordem: {data};{linha};{valor}
                complemento_geral = parametro.get("complemento", "")
                
                # Determina os complementos específicos baseados no tipo (débito/crédito)
                # Se não houver complemento específico, usa o complemento geral
                comp_vlr = complemento_geral
                comp_taxa = complemento_geral
                
                # Se for débito
                if "debito" in tipo_identificado or (usar_identificacao_por_coluna and tipo_linha == "debito"):
                    comp_vlr = parametro.get("complemento_debito") or complemento_geral
                    comp_taxa = parametro.get("complemento_debito_desconto") or complemento_geral
                # Se for crédito
                elif "credito" in tipo_identificado or (usar_identificacao_por_coluna and tipo_linha == "credito"):
                    comp_vlr = parametro.get("complemento_credito") or complemento_geral
                    comp_taxa = parametro.get("complemento_credito_desconto") or complemento_geral
                
                suffix_vlr = f";{comp_vlr}" if comp_vlr else ""
                suffix_taxa = f";{comp_taxa}" if comp_taxa else ""
                
                linha_valor = f"{data_formatada};{linha1};{valor_formatado}{suffix_vlr}"
                linha_taxa = f"{data_formatada};{linha2};{taxa_formatada}{suffix_taxa}"

                # Ordem de inserção: primeiro o valor (linha1), depois a taxa (linha2)
                linhas_saida.append(linha_valor)
                linhas_saida.append(linha_taxa)

            linha_atual += 1

            # Proteção contra loop infinito
            if linha_atual > ws.max_row + 100:
                break

        # Gera o conteúdo TXT
        # ⭐ CORREÇÃO: Usa \r\n (Windows) e adiciona linha vazia no final
        txt_content = "\r\n".join(linhas_saida) + "\r\n"
        
        return txt_content, None

    except Exception as e:
        import traceback
        erro_completo = f"Erro ao processar arquivo:\n{str(e)}\n\nDetalhes técnicos:\n{traceback.format_exc()}"
        return None, erro_completo


def process_file_worker(session_id, file_id, caminho_arquivo, nome_original):
    """Worker para processar arquivo em thread separada"""
    conn = get_conn()  # conexão própria da thread — sqlite3 não é thread-safe entre conexões compartilhadas
    try:
        # Processa o arquivo
        txt_content, erro = process_xlsx_to_txt(caminho_arquivo, nome_original)

        if erro:
            conn.execute(
                "UPDATE files SET status='error', error=? WHERE file_id=?", (erro, file_id)
            )
        else:
            # Salva o TXT gerado
            nome_base = os.path.splitext(nome_original)[0]
            nome_txt = f"{nome_base}.txt"
            txt_path = os.path.join(Config.OUTPUT_FOLDER, f"{session_id}_{file_id}_{nome_txt}")

            with open(txt_path, 'w', encoding='utf-8') as f:
                f.write(txt_content)

            conn.execute(
                "UPDATE files SET status='completed', txt_content=?, txt_filename=?, txt_path=? WHERE file_id=?",
                (txt_content, nome_txt, txt_path, file_id)
            )

    except Exception as e:
        import traceback
        erro_completo = f"Erro inesperado no worker:\n{str(e)}\n\n{traceback.format_exc()}"
        conn.execute(
            "UPDATE files SET status='error', error=? WHERE file_id=?", (erro_completo, file_id)
        )

    finally:
        conn.close()
        # Remove o arquivo XLSX temporário após o processamento
        if os.path.exists(caminho_arquivo):
            try:
                os.remove(caminho_arquivo)
            except:
                pass


@app.route('/')
def index():
    """Página principal"""
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_files():
    """Endpoint para upload de arquivos"""
    session_id = get_session_id()
    init_session_data(session_id)

    if 'files' not in request.files:
        return jsonify({'error': MESSAGES['upload_error']}), 400

    files = request.files.getlist('files')
    
    if not files or all(f.filename == '' for f in files):
        return jsonify({'error': MESSAGES['upload_error']}), 400

    # Limpa arquivos já processados antes de adicionar novos
    cleared_previous = cleanup_completed_files(session_id)

    uploaded_files = []
    conn = get_conn()
    try:
        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                file_id = str(uuid.uuid4())

                # Salva arquivo temporariamente
                filepath = os.path.join(Config.UPLOAD_FOLDER, f"{session_id}_{file_id}_{filename}")
                file.save(filepath)

                # Armazena informações do arquivo
                conn.execute(
                    "INSERT INTO files (file_id, session_id, filename, original_filename, filepath, status, upload_time) "
                    "VALUES (?, ?, ?, ?, ?, 'uploaded', ?)",
                    (file_id, session_id, filename, file.filename, filepath, datetime.now().isoformat())
                )

                uploaded_files.append({
                    'file_id': file_id,
                    'filename': file.filename,
                    'status': 'uploaded'
                })
    finally:
        conn.close()

    return jsonify({
        'message': MESSAGES['upload_success'].format(count=len(uploaded_files)),
        'files': uploaded_files,
        'cleared_previous': cleared_previous
    })


@app.route('/process', methods=['POST'])
def process_files():
    """Endpoint para processar arquivos"""
    session_id = get_session_id()
    init_session_data(session_id)

    try:
        ok, erro = try_start_processing(session_id)
    except sqlite3.OperationalError:
        return jsonify({'error': 'Sistema ocupado, tente novamente'}), 503

    if not ok:
        return jsonify({'error': erro}), 400

    # Inicia threads para processar cada arquivo
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT file_id, filepath, original_filename FROM files WHERE session_id=? AND status='uploaded'",
            (session_id,)
        ).fetchall()

        threads = []
        for file_id, filepath, original_filename in rows:
            conn.execute("UPDATE files SET status='processing' WHERE file_id=?", (file_id,))
            thread = threading.Thread(
                target=process_file_worker,
                args=(session_id, file_id, filepath, original_filename)
            )
            thread.start()
            threads.append(thread)
    finally:
        conn.close()

    # Aguarda todas as threads finalizarem
    try:
        for thread in threads:
            thread.join()
    finally:
        set_processing(session_id, 0)  # fix do bug 1.3: sempre reseta, exceção ou não

    return jsonify({'message': MESSAGES['processing_complete']})


@app.route('/status', methods=['GET'])
def get_status():
    """Endpoint para obter status do processamento"""
    session_id = get_session_id()
    init_session_data(session_id)

    conn = get_conn()
    try:
        processing_row = conn.execute(
            "SELECT processing FROM sessions WHERE session_id=?", (session_id,)
        ).fetchone()
        files = conn.execute(
            "SELECT file_id, original_filename, status, error FROM files WHERE session_id=?", (session_id,)
        ).fetchall()
    finally:
        conn.close()

    files_status = [
        {'file_id': f[0], 'filename': f[1], 'status': f[2], 'error': f[3]} for f in files
    ]

    return jsonify({
        'processing': bool(processing_row[0]) if processing_row else False,
        'files': files_status
    })


@app.route('/download/<file_id>', methods=['GET'])
def download_file(file_id):
    """Endpoint para download do arquivo TXT gerado"""
    session_id = get_session_id()
    init_session_data(session_id)

    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT status, original_filename, txt_path, txt_content, txt_filename FROM files "
            "WHERE file_id=? AND session_id=?",
            (file_id, session_id)
        ).fetchone()
    finally:
        conn.close()

    if row is None:
        return jsonify({'error': MESSAGES['file_not_found']}), 404

    status, original_filename, txt_path, txt_content, txt_filename = row

    if status != 'completed':
        return jsonify({'error': MESSAGES['file_not_processed']}), 400

    # Verifica se o arquivo TXT existe no disco
    if txt_path and os.path.exists(txt_path):
        # Envia o arquivo do disco
        return send_file(
            txt_path,
            as_attachment=True,
            download_name=txt_filename,
            mimetype='text/plain'
        )
    else:
        # Fallback: cria arquivo em memória (caso o arquivo no disco tenha sido removido)
        nome_base = os.path.splitext(original_filename)[0]
        nome_txt = f"{nome_base}.txt"

        buffer = io.BytesIO()
        buffer.write(txt_content.encode('utf-8'))
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=nome_txt,
            mimetype='text/plain'
        )


@app.route('/clear', methods=['POST'])
def clear_files():
    """Endpoint para limpar arquivos da sessão"""
    session_id = get_session_id()

    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT filepath, txt_path FROM files WHERE session_id=?", (session_id,)
        ).fetchall()
        for filepath, txt_path in rows:
            for p in (filepath, txt_path):
                if p and os.path.exists(p):
                    try:
                        os.remove(p)
                    except:
                        pass
        conn.execute("DELETE FROM files WHERE session_id=?", (session_id,))
        conn.execute("UPDATE sessions SET processing=0 WHERE session_id=?", (session_id,))
    finally:
        conn.close()

    return jsonify({'message': MESSAGES['clear_success']})


@app.route('/health', methods=['GET'])
def health_check():
    """Endpoint de health check"""
    # Conta arquivos nas pastas
    uploads_count = len([f for f in os.listdir(Config.UPLOAD_FOLDER) if f != '.gitkeep'])
    outputs_count = len([f for f in os.listdir(Config.OUTPUT_FOLDER) if f != '.gitkeep'])

    conn = get_conn()
    try:
        active_sessions = conn.execute("SELECT COUNT(*) FROM sessions").fetchone()[0]
    finally:
        conn.close()

    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'active_sessions': active_sessions,
        'uploads_folder_files': uploads_count,
        'outputs_folder_files': outputs_count
    })


@app.route('/debug/files', methods=['GET'])
def debug_files():
    """Endpoint para debug - lista arquivos nas pastas"""
    uploads_files = [f for f in os.listdir(Config.UPLOAD_FOLDER) if f != '.gitkeep']
    outputs_files = [f for f in os.listdir(Config.OUTPUT_FOLDER) if f != '.gitkeep']
    
    return jsonify({
        'uploads': {
            'count': len(uploads_files),
            'files': uploads_files
        },
        'outputs': {
            'count': len(outputs_files),
            'files': outputs_files
        }
    })


if __name__ == '__main__':
    app.run(debug=Config.DEBUG, host=Config.HOST, port=Config.PORT)